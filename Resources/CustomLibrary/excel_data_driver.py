import openpyxl
import datetime


def read_data(filepath):
    # Open the workbook
    workbook = openpyxl.load_workbook(filepath)
    # Select the sheet
    sheet = workbook.active
    # Get the header row values
    header_row = sheet[1]
    headers = [cell.value for cell in header_row]
    # Iterate through rows and create a dictionary for each row using the headers
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)
    return data

def write_data(data,filepath=None,sheetname=None):
    wb = openpyxl.Workbook()
    if sheetname is None:
        sheet_name = 'Sheet 1'
    else:
        sheet_name = sheetname
    if sheet_name not in wb.worksheets:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    # Write the headers to the worksheet
    headers = list(data[0].keys())
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write the data to the worksheet
    for row, row_data in enumerate(data, start=2):
        for col, col_data in enumerate(row_data.values(), start=1):
            ws.cell(row=row, column=col, value=col_data)

    if filepath is None:
        now = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename = f'data_{now}.xlsx'
        path='/Users/adusumilli/Documents/GitHub/Salesforce-Testing-RobotFramework/Resources/Data'
        wb.save(path+filename)
    else:
        wb.save(filepath)
    



